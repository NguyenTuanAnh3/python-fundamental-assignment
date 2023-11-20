import os
from datetime import datetime
import openpyxl
from openpyxl import workbook
from openpyxl import load_workbook

def create_excel(car_identity, leave_day, bills):
    path = f'bills/{car_identity}.xlsx'
    leave_day = datetime.strptime(leave_day,"%Y-%m-%d %H:%M").strftime("%Y-%m-%d")
    if not os.path.isdir('bills/'):
        os.mkdir('bills/')
    
    if not os.path.exists(path):
        wb = openpyxl.Workbook()
        wb.active.title = leave_day
        wb.save(path)

    wb = load_workbook(path)
    
    if leave_day not in wb.sheetnames:
        wb.create_sheet(leave_day)
    else:
        del wb[leave_day]
        wb.create_sheet(leave_day)

    ws = wb.active
    ws = wb[leave_day]

    format_excel(ws, bills)
    
    wb.save(path)

def format_excel(ws, bills):
    num_col = len([bill['date'] for bill in bills])
    create_title_yaxis(ws=ws)
    create_title_xaxis(ws=ws, bills=bills, num_col = num_col)
    filled_data(ws=ws, bills=bills, num_col = num_col)

def create_title_yaxis(ws):
    
    rows = ['Day', 'Day of Week', '08:00 - 16:59', '08:00 - 16:59', '17:00 - 23:59','23:59 - 7:59', 'Sub Total', 'Total']
    for row in range(1,9):
        ws.cell(row=row, column= 1, value=rows[row-1])
        if(row == 4):
            ws.merge_cells(start_row = row - 1, start_column= 1 , end_row= row, end_column= 1)

def create_title_xaxis(ws, bills, num_col):
    end_col = num_col * 2 + 1
    get_date = [bill['date'] for bill in bills]
    get_day_of_week = [bill['dayOfWeek'] for bill in bills]
    count_date = 0
    count_day_of_week = 0
    for row in range(1, 3):
        for col in range(2, end_col):
            if (col % 2 == 0):
                if(row % 2 != 0):
                    ws.cell(row=row, column=col, value= get_date[count_date])
                    count_date+=1
                else:
                    ws.cell(row=row, column= col, value= get_day_of_week[count_day_of_week])
                    count_day_of_week+=1

    for row in range(1,3):
        for col in range(2, end_col):
            if(col %2 == 0):    
                ws.merge_cells(start_row = row, start_column = col, end_column = col+1, end_row = row)

def filled_data(ws, bills, num_col):
    end_col = num_col * 2 + 2
    get_data_day_time = [ bill['08:00 - 16:59'] for bill in bills ]
    get_data_night_time = [ bill['17:00 - 23:59'] for bill in bills ]
    get_data_over_night = [ bill['00:00 - 07:59'] for bill in bills ]
    get_data_sub_total = [bill['sub_total'] for bill in bills]

    datas = []
    for index, data in enumerate(get_data_day_time):
        for i, d in enumerate(data):
            if i % 2 == 0:
                datas.append(f'{str(d["time"]) + " hours" if d["time"] != 0 else ""}')
                datas.append(f'{d["str_price_per_hour"]}')

    for index, data in enumerate(get_data_day_time):
        for i, d in enumerate(data):
            if i % 2 != 0:
                datas.append(f'{str(d["time"]) + " hours" if d["time"] != 0 else ""}')
                datas.append(f'{d["str_price_per_hour"]}')
    
    for index, data in enumerate(get_data_night_time):
        datas.append(f'{str(data[0]["time"]) + " hours" if data[0]["time"] != 0 else ""}')
        datas.append(f'{data[0]["str_price_per_hour"]}')

    for index, data in enumerate(get_data_over_night):
        datas.append(f'{str(data[0]["time"]) if data[0]["time"] != 0 else ""}')
        datas.append(f'{data[0]["str_price_per_hour"]}')


    count = 0
    for row in range(3,7):
        for col in range(2, end_col):
            ws.cell(row=row, column= col, value= datas[count])
            count+=1
    
    count = 0
    for col in range(2, end_col):
        if(col % 2 == 0 ):
            ws.cell(row= 7, column= col, value= f'${"{0:.2f}".format(get_data_sub_total[count])}')
            count+=1

    total = sum(get_data_sub_total)
    ws.cell(row= 8, column= 2, value= f'${"{0:.2f}".format(total)}')
    